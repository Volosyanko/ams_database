# install packages
import json
import os
from io import BytesIO
from tempfile import NamedTemporaryFile

import flask
from flask import (
    Flask,
    render_template,
    request,
    jsonify,
    send_file,
    Response,
    redirect,
    url_for,
)
from flask_mysqldb import MySQL, MySQLdb
from openpyxl import load_workbook, Workbook

# create an application object
app = Flask(__name__)

# connect to SQL server
app.config["MYSQL_HOST"] = "localhost"
app.config["MYSQL_USER"] = "root"
app.config["MYSQL_PASSWORD"] = "password"
app.config["MYSQL_DB"] = "aerosol_ms_database"
app.config["MYSQL_PORT"] = 3306
mysql = MySQL(app)

# assign all column headers from excel file
# these must match (in order) with the column headers found on your SQL database
HEADERS = [
    "CompoundName",
    "ExperimentType",
    "Compound Type",
    "ChemicalFormula",
    "Category",
    "CAS",
    "Ionization",
    "ResponseThreshold",
    "Internal Standard",
    "Internal Standard Concentration",
    "PrecursorMass",
    "ExtractedMass",
    "Adduct",
    "Polarity",
    "ChargeState",
    "RT",
    "Window",
    "CollisionEnergy",
    "Lens",
    "EnergyRamp",
    "Confirm Precursor",
    "Confirm Extracted",
    "Confirm Energy",
    "Target Ratio",
    "Window Type",
    "Ratio Window",
    "Ion Coelution",
    "Confirm Precursor",
    "Confirm Extracted",
    "Confirm Energy",
    "Target Ratio",
    "Window Type",
    "Ratio Window",
    "Ion Coelution",
    "Confirm Precursor",
    "Confirm Extracted",
    "Confirm Energy",
    "Target Ratio",
    "Window Type",
    "Ratio Window",
    "Ion Coelution",
    "Confirm Precursor",
    "Confirm Extracted",
    "Confirm Energy",
    "Target Ratio",
    "Window Type",
    "Ratio Window",
    "Ion Coelution",
    "Confirm Precursor",
    "Confirm Extracted",
    "Confirm Energy",
    "Target Ratio",
    "Window Type",
    "Ratio Window",
    "Ion Coelution",
    "Fragment",
    "Fragment",
    "Fragment",
    "Fragment",
    "Fragment",
]

# assign columns headers you would like displayed in table on webpage
SEARCH_COLUMNS = [
    "CompoundName",
    "ChemicalFormula",
    "Ionization",
    "ExtractedMass",
    "RT",
    "CollisionEnergy",
    "ConfirmExtracted1",
    "ConfirmExtracted2",
    "ConfirmPrecursor3",
    "ConfirmExtracted3",
    "ConfirmEnergy3",
    "ConfirmExtracted4",
    "Fragment1",
    "Fragment2",
    "Fragment3",
    "Fragment4",
    "Fragment5",
]


# code for main page
@app.route("/")
def index():
    return render_template("index.html")


# code for upload page
@app.route("/upload")
def upload():
    return render_template("upload.html")


# code for import feature
@app.route("/parse", methods=["POST"])
def parse():
    files = flask.request.files.getlist("file")
    cur = mysql.connection.cursor()
    for file in files:
        workbook = load_workbook(BytesIO(file.stream.read()))
        worksheet = workbook.active
        row_number = 1
        for row_number, row in enumerate(
            worksheet.iter_rows(max_row=3, values_only=True), start=1
        ):
            if row[0] in (
                "CompoundName",
                "ExperimentType",
                "Compound Type",
                "ChemicalFormula",
                "Fragment",
                "Category",
                "CAS",
                "Ionization",
                "ResponseThreshold",
                "Internal Standard",
                "Concentration",
                "PrecursorMass",
            ):
                break
            row_number += 1
        for row in worksheet.iter_rows(min_row=row_number + 1, values_only=True):
            query = f"""
            INSERT INTO newnitromix1 (CompoundName, ExperimentType, CompoundType, ChemicalFormula, Category, CAS,                    
    Ionization, ResponseThreshold, InternalStandard, InternalStandardConcentration, PrecursorMass, ExtractedMass, 
    Adduct, Polarity,ChargeState, RT, `Window`, CollisionEnergy, Lens, EnergyRamp, ConfirmPrecursor1, ConfirmExtracted1, 
    ConfirmEnergy1,TargetRatio1, WindowType1, RatioWindow1, IonCoelution1, ConfirmPrecursor2, ConfirmExtracted2,
     ConfirmEnergy2, TargetRatio2, WindowType2, RatioWindow2, IonCoelution2, ConfirmPrecursor3, ConfirmExtracted3,
      ConfirmEnergy3, TargetRatio3, WindowType3,RatioWindow3, IonCoelution3, ConfirmPrecursor4, ConfirmExtracted4,
       ConfirmEnergy4, TargetRatio4,WindowType4, RatioWindow4, IonCoelution4, ConfirmPrecursor5,
        ConfirmExtracted5, ConfirmEnergy5, TargetRatio5, WindowType5, RatioWindow5,IonCoelution5,Fragment1, Fragment2,
        Fragment3,Fragment4, Fragment5)
     VALUES ({'%s, ' * (len(row) - 1)} %s) 
            """
            cur.execute(query, row)
            mysql.connection.commit()
    return redirect(url_for("index"))


# code for export feature
@app.route("/download", methods=["POST"])
def download():
    data = {i["name"]: i["value"] for i in json.loads(request.get_data().decode())}
    condition = ""
    cur = mysql.connection.cursor()
    if data.get("search_text"):
        search_word = data.get("search_text")
        query = """
        SELECT * from newnitromix1 
        WHERE CompoundName LIKE %s OR ChemicalFormula LIKE %s OR Ionization LIKE %s OR
            ExtractedMass LIKE %s OR RT LIKE %s OR CollisionEnergy LIKE %s OR ConfirmExtracted1 LIKE %s
            OR ConfirmExtracted2 LIKE %s OR ConfirmPrecursor3 LIKE %s OR ConfirmExtracted3 LIKE %s
            OR ConfirmEnergy3 LIKE %s OR ConfirmExtracted4 LIKE %s OR Fragment1 LIKE %s
            OR Fragment2 LIKE %s OR Fragment3 LIKE %s OR Fragment4 LIKE %s
            OR Fragment5 LIKE %s
        """
        arg = "%" + search_word + "%"
        args = [arg for _ in SEARCH_COLUMNS]
        cur.execute(query, args)
    else:
        args = []
        if data.get("compound_name_search"):
            condition += "WHERE CompoundName LIKE %s"
            args.append("%" + data.get("compound_name_search") + "%")
        if data.get("chemical_formula_search"):
            condition += (
                "WHERE ChemicalFormula LIKE %s"
                if not condition
                else " OR ChemicalFormula LIKE %s"
            )
            args.append("%" + data.get("chemical_formula_search") + "%")
        if data.get("ionization_search"):
            condition += (
                "WHERE Ionization LIKE %s"
                if not condition
                else " OR Ionization LIKE %s"
            )
            args.append("%" + data.get("ionization_search") + "%")
        if data.get("extracted_mass_search"):
            condition += (
                "WHERE ExtractedMass LIKE %s"
                if not condition
                else " OR ExtractedMass LIKE %s"
            )
            args.append("%" + data.get("extracted_mass_search") + "%")
        if data.get("rt_search"):
            condition += "WHERE RT LIKE %s" if not condition else " OR RT LIKE %s"
            args.append("%" + data.get("rt_search") + "%")
        if data.get("collision_energy_search"):
            condition += (
                "WHERE CollisionEnergy LIKE %s"
                if not condition
                else " OR CollisionEnergy LIKE %s"
            )
            args.append("%" + data.get("collision_energy_search") + "%")
        if data.get("confirm_extracted1_search"):
            condition += (
                "WHERE ConfirmExtracted1 LIKE %s"
                if not condition
                else " OR ConfirmExtracted1 LIKE %s"
            )
            args.append("%" + data.get("confirm_extracted1_search") + "%")
        if data.get("confirm_extracted2_search"):
            condition += (
                "WHERE ConfirmExtracted2 LIKE %s"
                if not condition
                else " OR ConfirmExtracted2 LIKE %s"
            )
            args.append("%" + data.get("confirm_extracted2_search") + "%")
        if data.get("confirm_extracted3_search"):
            condition += (
                "WHERE ConfirmExtracted3 LIKE %s"
                if not condition
                else " OR ConfirmExtracted3 LIKE %s"
            )
            args.append("%" + data.get("confirm_extracted3_search") + "%")
        if data.get("confirm_precursor3_search"):
            condition += (
                "WHERE ConfirmPrecursor3 LIKE %s"
                if not condition
                else " OR ConfirmPrecursor3 LIKE %s"
            )
            args.append("%" + data.get("confirm_precursor3_search") + "%")
        if data.get("confirm_extracted4_search"):
            condition += (
                "WHERE ConfirmExtracted4 LIKE %s"
                if not condition
                else " OR ConfirmExtracted4 LIKE %s"
            )
            args.append("%" + data.get("confirm_extracted4_search") + "%")
        if data.get("confirm_energy3_search"):
            condition += (
                "WHERE ConfirmEnergy3 LIKE %s"
                if not condition
                else " OR ConfirmEnergy3 LIKE %s"
            )
            args.append("%" + data.get("confirm_energy3_search") + "%")
        if data.get("confirm_energy3_search"):
            condition += (
                "WHERE ConfirmEnergy3 LIKE %s"
                if not condition
                else " OR ConfirmEnergy3 LIKE %s"
            )
            args.append("%" + data.get("confirm_energy3_search") + "%")
        if data.get("fragment1_search"):
            condition += (
                "WHERE Fragment1 LIKE %s" if not condition else " OR Fragment1 LIKE %s"
            )
            args.append("%" + data.get("fragment1_search") + "%")
        if data.get("fragment2_search"):
            condition += (
                "WHERE Fragment2 LIKE %s" if not condition else " OR Fragment2 LIKE %s"
            )
            args.append("%" + data.get("fragment2_search") + "%")
        if data.get("fragment3_search"):
            condition += (
                "WHERE Fragment3 LIKE %s" if not condition else " OR Fragment3 LIKE %s"
            )
            args.append("%" + data.get("fragment3_search") + "%")
        if data.get("fragment4_search"):
            condition += (
                "WHERE Fragment4 LIKE %s" if not condition else " OR Fragment4 LIKE %s"
            )
            args.append("%" + data.get("fragment4_search") + "%")
        if data.get("fragment5_search"):
            condition += (
                "WHERE Fragment5 LIKE %s" if not condition else " OR Fragment5 LIKE %s"
            )
            args.append("%" + data.get("fragment5_search") + "%")
        query = f"""
        SELECT * from newnitromix1
        {condition}
        """
        cur.execute(query, args)
    data = cur.fetchall()
    workbook = Workbook()
    worksheet = workbook.active
    tmp = NamedTemporaryFile(delete=False)
    worksheet.append(HEADERS)
    for row in data:
        worksheet.append(row)
    workbook.save(tmp.name)
    tmp.seek(0)
    stream = tmp.read()
    tmp.close()
    os.remove(tmp.name)
    return Response(
        stream,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-disposition": "attachment; filename=sample.xlsx"},
    )


# code to retrieve data from search all row
@app.route("/allsearch", methods=["POST", "GET"])
def allsearch():
    cur = mysql.connection.cursor(MySQLdb.cursors.DictCursor)
    if request.method == "POST":
        search_word = request.form["query"]
        print(search_word)
        if search_word == "":
            query = "SELECT * from newnitromix1"
            cur.execute(query)
            numrows = int(cur.rowcount)
            newnitromix1 = cur.fetchall()
        else:
            query = """SELECT * from newnitromix1
            WHERE CompoundName LIKE '%{}%' OR ChemicalFormula LIKE '%{}%' OR Ionization LIKE '%{}%' OR
            ExtractedMass LIKE '%{}%' OR RT LIKE '%{}%' OR CollisionEnergy LIKE '%{}%' OR ConfirmExtracted1 LIKE '%{}%'
            OR ConfirmExtracted2 LIKE '%{}%' OR ConfirmPrecursor3 LIKE '%{}%' OR ConfirmExtracted3 LIKE '%{}%'
            OR ConfirmEnergy3 LIKE '%{}%' OR ConfirmExtracted4 LIKE '%{}%' OR Fragment1 LIKE '%{}%'
            OR Fragment2 LIKE '%{}%' OR Fragment3 LIKE '%{}%' OR Fragment4 LIKE '%{}%'
            OR Fragment5 LIKE '%{}%'""".format(
                search_word,
                search_word,
                search_word,
                search_word,
                search_word,
                search_word,
                search_word,
                search_word,
                search_word,
                search_word,
                search_word,
                search_word,
                search_word,
                search_word,
                search_word,
                search_word,
                search_word,
            )
            cur.execute(query)
            numrows = int(cur.rowcount)
            newnitromix1 = cur.fetchall()
            print(numrows)
    return jsonify(
        {
            "htmlresponse": render_template(
                "response.html", newnitromix1=newnitromix1, numrows=numrows
            )
        }
    )


if __name__ == "__main__":
    app.run(debug=True)
